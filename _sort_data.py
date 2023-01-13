from _clean_data import load_file
from openpyxl.styles import Alignment
from operator import itemgetter


# constants used throughout data cleaning
NUM_CLASS_LIST_COLS = 11
NUM_ACTIVITY_HEADER_COLS = 7
NUM_MAX_STUDENTS = 32

def sort_class_data(sort_method, file_path, class_list, assessment_data):
    """
    Takes clean_sort method variable input by user to determine how to clean_sort organize the spreadsheet data; the user can switch
    between alphabetical order by last name and their Google ID (which is also their first name). Data is cleared from
    the spreadsheet and re-written based on the clean_sort method selected.
    :param sort_method: numeric input (1/2) for last name v. Google ID
    :param data: brings in data from appropriate document and respective class list sheet and assessment data sheet
    :param file_path: path to where the file is located
    :return: sorted data based on last name or Google ID / first name in array format
    """

    # open Excel File and assign sheets to variables
    wb = load_file(full_file_path=file_path)
    assessment_sheet = wb["Tracking"]
    profile_sheet = wb["Student_List"]

    # assign data to specific recognizable variables
    #class_list = data["class list"]
    #assessment_data = data["assessment data"]

    # combine the student list and assessment data sheet data into a single array to perform clean_sort
    combined_data_sets = []
    for student_row in range(0, len(class_list)):
        student_data = []
        for col in range(0, NUM_CLASS_LIST_COLS):
            student_data.append(class_list[student_row][col])
        student_data.append("|*|")
        for activity in range(0, len(assessment_data)):
            student_data.append(assessment_data[activity][NUM_ACTIVITY_HEADER_COLS + student_row])
        combined_data_sets.append(student_data)

    # perform the clean_sort
    sorted_list = ""
    if sort_method == "alpha":
        # by last name
        sorted_list = sorted(combined_data_sets, key=itemgetter(1))
    elif sort_method == "googleID":
        # by GoogleID / first name
        sorted_list = sorted(combined_data_sets, key=itemgetter(4))

    # clean profile sheet (remove cell contents on the student list sheet)
    for row in range(2, NUM_MAX_STUDENTS+1):
        for col in range(1, NUM_CLASS_LIST_COLS+1):
            cell = profile_sheet.cell(row, col)
            value = None
            cell.value = value
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       text_rotation=0,
                                       wrap_text=True)

    # write sorted data to the profile / class list sheet
    starting_cell = [2, 1]  # row = 2, col = 1 ('A2')
    for row in range(0, len(sorted_list)):
        for col in range(0, NUM_CLASS_LIST_COLS):
            cell = profile_sheet.cell(starting_cell[0] + row, starting_cell[1] + col)
            value = sorted_list[row][col]
            cell.value = value
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       text_rotation=0,
                                       wrap_text=True)

    # clear assessment sheet header row
    for c in range(0, 32):
        cell = assessment_sheet.cell(1, 8 + c)
        cell.value = None
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center',
                                   text_rotation=80,
                                   wrap_text=True)

    # clear the assessment tracking sheet, only student achievement data
    for row in range(0, len(assessment_data)):
        for col in range(0, len(class_list)):
            cell = assessment_sheet.cell(2+row, 8+col)
            value = None
            cell.value = value
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       text_rotation=0,
                                       wrap_text=True)

    # reassign students to header row after clean_sort
    cell = assessment_sheet.cell(1, 8)
    cell.value = '=TRANSPOSE(Student_List!E2:E32)'
    assessment_sheet.formula_attributes['H1'] = {'t': 'array', 'ref': "H1:H1"}
    # TODO: above line helps to write the formula in a format in which it is read, without the above line the formula
    #  gets written with @ symbols ( e.g. =@TRANSPOSE(@Student_List!E2:E32) ); with this line the formula is written
    #  as an array formula ( e.g. ={TRANSPOSE(Student_List!E2:E32) ); need to determine how to use this line to have it
    #  write straight so that the user does not need to correct the formula

    # add assessment data back in to 'tracking sheet' after clean_sort
    starting_cell = [2, 8]  # row = 2, col = 8 ('H2')
    for col in range(0, len(sorted_list)):
        for row in range(0, len(assessment_data)):
            cell = assessment_sheet.cell(starting_cell[0] + row, starting_cell[1] + col)
            # 12 >>> pass over student list info + "|*|" to get to achievement data
            value = sorted_list[col][row+12]
            cell.value = value
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       text_rotation=0,
                                       wrap_text=True)

    wb.save(filename=file_path)

    return sorted_list